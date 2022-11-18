class bcolors: # - Adds the terminal colors
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

#header
print(bcolors.HEADER + bcolors.OKBLUE + bcolors.BOLD + '==================================================== SPEEDTEST SHEETS ====================================================' + bcolors.ENDC)

#imports
import speedtest as ss #speedtest.net
import time #used for time & date
from pathlib import Path #checks to make sure theres not multiple sheets
from openpyxl import Workbook, load_workbook as excel #How the spreadsheet is made and edited (openpyxl)
import sys #used to exit the code without error message

#the file name (name it to whatever just make sure it ends with .xlsx)
FILE_NAME = 'SHEET.xlsx'

#making sure its ready to edit a new sheet
path = Path(FILE_NAME)
if path.is_file():
    print(bcolors.FAIL + bcolors.BOLD + 'Please remove the old spreadsheet from the current directory.')
    sys.exit("")
else:
    print('')

#variables for, speedtest & dates and time
st = ss.Speedtest()
local_time = time.localtime()
systemtime = time.strftime('%a, %d %b %Y %H:%M:%S', local_time)
date = '%b %d %Y'
time_ = '%H:%M:%S'
system_date = time.strftime(date, local_time)
system_time = time.strftime(time_, local_time)

#enabling openpyxl
wb = Workbook()
ws = wb.active
ws.title = "SPEEDTEST SHEET"

#getting the users integer input for how many x seconds it should check the speeds
valid = False
while not valid:
    try:
        print(bcolors.OKGREEN + '------------------------------------------------ Ready! ------------------------------------------------' + bcolors.ENDC)
        print(bcolors.WARNING + 'How many x seconds should it wait to check & log again? !IN SECONDS!' + bcolors.ENDC) 
        print(bcolors.WARNING + 'NOTE: the speedtest takes 10-20 seconds onece activated every x seconds') #Every time the set amount of seconds is reached it takes a bit of time for the speedtest to load just like on a speedtest website (speedtest.net)
        print(f'{bcolors.OKCYAN}Time converter: {bcolors.UNDERLINE}https://etahn.ml/time-converter' + bcolors.ENDC)
        looptime = int(input(bcolors.OKBLUE + bcolors.BOLD + 'Input: x= ' + bcolors.ENDC)) #Integer Input
        valid = True
    except ValueError:
        print(bcolors.FAIL + 'Please only input digits' + bcolors.ENDC)
print(bcolors.OKGREEN + f'Started! Checking every {looptime} seconds.' + bcolors.ENDC) #Telling the user its starting.

# The loop:
while True:
    #again the dates and time being refreshed every round it loops.
    local_time = time.localtime()
    systemtime = time.strftime('%a, %d %b %Y %H:%M:%S', local_time)
    date = '%b %d %Y'
    time_ = '%H:%M:%S'
    system_date = time.strftime(date, local_time)
    system_time = time.strftime(time_, local_time)
    
    
    download_speed = st.download()
    upload_speed = st.upload()
    download = '{:5.2f} Mb'.format(download_speed/(1024*1024))
    upload = '{:5.2f} Mb'.format(upload_speed/(1024*1024))
    
    print(bcolors.OKCYAN + '-----------------------------------')
    print(systemtime)
    print('Download Speed: {:5.2f} Mb'.format(download_speed/(1024*1024) ))
    print('Upload Speed: {:5.2f} Mb'.format(upload_speed/(1024*1024) ))
    bcolors.ENDC
    
    #again the dates and time being refreshed every round it loops. 2nd so it keeps the variables updated
    local_time = time.localtime()
    systemtime = time.strftime('%a, %d %b %Y %H:%M:%S', local_time)
    date = '%b %d %Y'
    time_ = '%H:%M:%S'
    system_date = time.strftime(date, local_time)
    system_time = time.strftime(time_, local_time)
    
    #checks to see if a sheet with its date already exists if not goes to else: and creates a new one with the date
    if system_date in wb.sheetnames: #
        print('') #nothing
    else:
        wb.create_sheet(system_date) #creates the sheet with the current date
        ws = wb[system_date] #Top left (a1) adding the date
        ws.append([system_date, 'Download / MB', 'Upload / MB']) #Top heard (row 1)
    
    ws.append([system_time, download, upload]) #Every time it loops adds a new line with the time of day and speeds
    wb.save(FILE_NAME) #Creates if not already made and saves the file
    time.sleep(looptime) #How long before it loops again (set with the integer value)
# https://github.com/etahn-git/SPEEDTEST-SHEETS